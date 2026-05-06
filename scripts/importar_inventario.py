# -*- coding: utf-8 -*-
"""
Importa el inventario consolidado al ERP.

Lee data/inventario/inventario-consolidado.xlsx y crea:

1. Productos en catalogo_productos (con moneda_compra='USD' y costo_unitario_usd)
2. Stock en inventario (almacen Contenedores - compartido del grupo)
3. Movimientos en inventario_movimientos:
   - tipo='carga_inicial_gasto'
   - registro_contable_pendiente=TRUE
   - empresa_pago_id=NULL (control llenará después según mapeo de color)
   - motivo y observaciones describen el origen

POLÍTICA CONTABLE: ver Configuración del Excel. Estos productos fueron pagados
como GASTO en cada empresa pagadora. Aquí solo registramos el stock físico
para visibilidad operativa, sin afectar contabilidad.

Uso:
    python scripts/importar_inventario.py [--dry-run]
"""

import json
import os
import sys
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# --- Config ---
SRC = Path(__file__).parent.parent / "data" / "inventario" / "inventario-consolidado.xlsx"
ALMACEN_DESTINO_CODIGO = "CONTENEDORES"
TC_ACTUAL = 17.2530
DRY_RUN = "--dry-run" in sys.argv

SUPABASE_URL = "https://dtmcqjtqykbkapzebbik.supabase.co"
SERVICE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6"
    "ImR0bWNxanRxeWtia2FwemViYmlrIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6"
    "MTc3NzU5MjM4MCwiZXhwIjoyMDkzMTY4MzgwfQ.-0TfmY0JaZTSn62jqtcmeroNeLX99"
    "Soa654OVbO1hwY"
)

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
}


def http(
    method: str,
    path: str,
    body: Any | None = None,
    params: dict | None = None,
    prefer: str | None = None,
) -> tuple[int, str]:
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        sep = "?" if "?" not in url else "&"
        url += sep + "&".join(f"{k}={v}" for k, v in params.items())
    h = dict(HEADERS)
    if prefer:
        h["Prefer"] = prefer
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method, headers=h)
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.status, r.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8")


def encontrar_usuario_capturador() -> str:
    """Devuelve UUID del CEO para usar como capturado_por."""
    s, body = http("GET", "usuarios_empresas?select=usuario_id&rol=eq.ceo&limit=1")
    rows = json.loads(body)
    if not rows:
        raise SystemExit("No se encontró un usuario CEO para capturado_por.")
    return rows[0]["usuario_id"]


def encontrar_almacen() -> tuple[str, str]:
    s, body = http(
        "GET",
        f"almacenes?select=id,empresa_id,codigo&codigo=eq.{ALMACEN_DESTINO_CODIGO}&limit=1",
    )
    rows = json.loads(body)
    if not rows:
        raise SystemExit(f"No existe almacén {ALMACEN_DESTINO_CODIGO}.")
    return rows[0]["id"], rows[0].get("empresa_id")


# ============================================================================
# Lectura del Excel consolidado
# ============================================================================


def leer_estructura_pv(wb: openpyxl.Workbook) -> list[dict]:
    """SKUs consolidados de estructura PV."""
    ws = wb["Estructura PV"]
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        codigo, descripcion, stock, n_lotes, n_emp, precio_usd, *_ = row
        if not codigo:
            continue
        items.append(
            {
                "codigo": str(codigo),
                "nombre": str(descripcion)[:200],
                "descripcion": str(descripcion),
                "stock": int(stock),
                "precio_usd": float(precio_usd or 0),
                "n_lotes": int(n_lotes or 0),
                "categoria": categorizar_estructura(str(codigo), str(descripcion)),
                "marca": "NEXT" if str(codigo).startswith("NXT-") else extraer_marca(str(descripcion)),
                "unidad_medida": "pieza",
            }
        )
    return items


def categorizar_estructura(codigo: str, descripcion: str) -> str:
    """Devuelve un valor del enum categoria_inventario."""
    c = codigo.upper()
    d = descripcion.upper()
    if "PANEL" in d or "JA-M66" in c:
        return "panel_solar"
    if "CABLE" in d or c.startswith("PV-C"):
        return "cable"
    if "CONECTOR" in d or "MC4" in d:
        return "herraje"
    return "estructura"


def categorizar_inversor_enum(categoria_excel: str) -> str:
    """Convierte la categoría del Excel al enum categoria_inventario."""
    c = (categoria_excel or "").lower()
    if "micro" in c:
        return "inversor"  # microinversores también son inversores en el enum
    if "monitoreo" in c:
        return "monitoreo"
    if "accesorio" in c:
        return "herraje"
    if "aislado" in c:
        return "inversor"
    return "inversor"


def extraer_marca(descripcion: str) -> str:
    d = descripcion.upper()
    if "JA SOLAR" in d or "JA-M66" in d:
        return "JA SOLAR"
    if "NEXT" in d:
        return "NEXT"
    return "GENERICO"


def leer_inversores(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Inversores"]
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        (
            codigo,
            marca,
            descripcion,
            categoria,
            capacidad,
            ini,
            ent,
            sal,
            stk,
            ppw,
            precio_usd,
            *_,
        ) = row
        if not codigo:
            continue
        stock_actual = int(ini or 0) + int(ent or 0) - int(sal or 0)
        capacidad_w = None
        if isinstance(capacidad, str) and "kW" in capacidad:
            try:
                capacidad_w = float(capacidad.replace(" kW", "").strip()) * 1000
            except ValueError:
                pass
        items.append(
            {
                "codigo": str(codigo),
                "nombre": str(descripcion)[:200],
                "descripcion": str(descripcion),
                "stock": stock_actual,
                "stock_inicial": int(ini or 0),
                "entrada": int(ent or 0),
                "salida": int(sal or 0),
                "precio_usd": float(precio_usd or 0),
                "categoria": categorizar_inversor_enum(str(categoria or "inversor")),
                "marca": str(marca),
                "capacidad_w": capacidad_w,
                "unidad_medida": "pieza",
            }
        )
    return items


def leer_cables(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Cables FV"]
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        codigo, descripcion, unidad, ini, ent, sal, stk, precio_usd, *_ = row
        if not codigo:
            continue
        stock_actual = int(ini or 0) + int(ent or 0) - int(sal or 0)
        items.append(
            {
                "codigo": str(codigo),
                "nombre": str(descripcion)[:200],
                "descripcion": str(descripcion),
                "stock": stock_actual,
                "stock_inicial": int(ini or 0),
                "entrada": int(ent or 0),
                "salida": int(sal or 0),
                "precio_usd": float(precio_usd or 0),
                "categoria": "cable",
                "marca": "GENERICO",
                "unidad_medida": str(unidad or "metros"),
            }
        )
    return items


def leer_lotes_detalle(wb: openpyxl.Workbook) -> list[dict]:
    """Lotes originales para crear inventario_movimientos con audit trail."""
    ws = wb["Lotes detalle"]
    lotes = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        (
            codigo,
            descripcion,
            cantidad,
            cant_orig,
            correccion,
            precio_usd,
            color,
            color_label,
            *_,
        ) = row
        if not codigo or not cantidad:
            continue
        lotes.append(
            {
                "codigo": str(codigo),
                "cantidad": int(cantidad),
                "cantidad_original": cant_orig if isinstance(cant_orig, int) else None,
                "correccion_motivo": str(correccion) if correccion else None,
                "precio_usd": float(precio_usd or 0),
                "color": str(color or ""),
                "color_label": str(color_label or ""),
            }
        )
    return lotes


# ============================================================================
# Inserción en BD
# ============================================================================


def upsert_producto(item: dict) -> str | None:
    """Upsert por codigo. Retorna producto_id."""
    payload = {
        "codigo": item["codigo"],
        "nombre": item["nombre"],
        "descripcion": item["descripcion"],
        "marca": item.get("marca"),
        "categoria": item.get("categoria"),
        "unidad_medida": item.get("unidad_medida", "pieza"),
        "moneda_compra": "USD",
        "costo_unitario_usd": round(item["precio_usd"], 4),
        "costo_promedio": round(item["precio_usd"] * TC_ACTUAL, 2),
        "tc_compra_referencia": TC_ACTUAL,
        "fecha_costo_actualizado": datetime.now().isoformat(),
        "activo": True,
    }
    if item.get("capacidad_w"):
        payload["capacidad"] = item["capacidad_w"]
        payload["unidad_capacidad"] = "W"

    if DRY_RUN:
        return "DRY-RUN-ID"

    s, body = http(
        "POST",
        "catalogo_productos",
        body=payload,
        prefer="resolution=merge-duplicates,return=representation",
        params={"on_conflict": "codigo"},
    )
    if s not in (200, 201):
        print(f"    ERROR upsert producto {item['codigo']}: {s} {body[:200]}")
        return None
    rows = json.loads(body) if body else []
    return rows[0]["id"] if rows else None


def upsert_inventario(almacen_id: str, producto_id: str, stock: int) -> bool:
    if DRY_RUN:
        return True
    payload = {
        "almacen_id": almacen_id,
        "producto_id": producto_id,
        "stock": stock,
        "ultima_entrada": datetime.now().isoformat(),
    }
    s, body = http(
        "POST",
        "inventario",
        body=payload,
        prefer="resolution=merge-duplicates,return=minimal",
        params={"on_conflict": "almacen_id,producto_id"},
    )
    if s not in (200, 201, 204):
        print(f"    ERROR upsert inventario: {s} {body[:200]}")
        return False
    return True


def insertar_movimiento(
    almacen_id: str,
    producto_id: str,
    cantidad: int,
    capturado_por: str,
    motivo: str,
    observaciones: str,
    tipo: str = "carga_inicial_gasto",
    costo_unitario_mxn: float | None = None,
    registro_contable_pendiente: bool = True,
) -> bool:
    if DRY_RUN:
        return True
    if cantidad <= 0:
        return True  # nada que insertar
    payload: dict = {
        "almacen_id": almacen_id,
        "producto_id": producto_id,
        "tipo": tipo,
        "cantidad": cantidad,
        "capturado_por": capturado_por,
        "motivo": motivo[:500],
        "observaciones": observaciones[:1000],
        "registro_contable_pendiente": registro_contable_pendiente,
    }
    if costo_unitario_mxn is not None:
        payload["costo_unitario"] = round(costo_unitario_mxn, 4)
    s, body = http("POST", "inventario_movimientos", body=payload, prefer="return=minimal")
    if s not in (200, 201, 204):
        print(f"    ERROR insertar movimiento ({tipo}): {s} {body[:200]}")
        return False
    return True


# ============================================================================
# Main
# ============================================================================


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"No existe {SRC}")

    print(f"Leyendo {SRC}...")
    wb = openpyxl.load_workbook(SRC, data_only=False)

    estructura = leer_estructura_pv(wb)
    inversores = leer_inversores(wb)
    cables = leer_cables(wb)
    lotes = leer_lotes_detalle(wb)

    print(f"  Estructura PV : {len(estructura)} SKUs")
    print(f"  Inversores    : {len(inversores)} SKUs")
    print(f"  Cables FV     : {len(cables)} SKUs")
    print(f"  Lotes detalle : {len(lotes)} compras originales")

    if DRY_RUN:
        print("\n*** DRY-RUN — no se hará ninguna escritura ***")

    print("\nResolviendo almacén destino y usuario capturador...")
    capturado_por = encontrar_usuario_capturador()
    almacen_id, _ = encontrar_almacen()
    print(f"  Capturado por : {capturado_por}")
    print(f"  Almacén destino: {almacen_id} ({ALMACEN_DESTINO_CODIGO})")

    print("\n=== 1. Upsert catalogo_productos ===")
    todos_items = estructura + inversores + cables
    productos_ids: dict[str, str] = {}
    for item in todos_items:
        pid = upsert_producto(item)
        if pid:
            productos_ids[item["codigo"]] = pid
    print(f"  {len(productos_ids)} productos en catálogo.")

    if DRY_RUN:
        print("\n=== DRY-RUN: simulando inserción de stock + movimientos ===")
        return

    print("\n=== 2. Upsert inventario (stock por SKU) ===")
    n_inv = 0
    for item in todos_items:
        pid = productos_ids.get(item["codigo"])
        if not pid:
            continue
        if item["stock"] <= 0:
            continue
        if upsert_inventario(almacen_id, pid, item["stock"]):
            n_inv += 1
    print(f"  {n_inv} registros de stock.")

    print("\n=== 3. Movimientos por lote (audit trail) ===")
    # Estructura: una entrada por cada lote original (39 lotes en Hoja1)
    n_mov = 0
    for lote in lotes:
        pid = productos_ids.get(lote["codigo"])
        if not pid:
            continue
        motivo = (
            f"Carga inicial - origen color {lote['color']} ({lote['color_label']})"
            + (
                f" — corrección: {lote['correccion_motivo']}"
                if lote.get("correccion_motivo")
                else ""
            )
        )
        observaciones = (
            f"Compra original con precio {lote['precio_usd']:.4f} USD/u. "
            f"Pagado como GASTO por la empresa con color {lote['color']}. "
            f"Pendiente reconciliación contable con control."
        )
        if insertar_movimiento(
            almacen_id, pid, lote["cantidad"], capturado_por, motivo, observaciones
        ):
            n_mov += 1
    print(f"  {n_mov} movimientos de lotes Estructura PV.")

    # Inversores: una entrada por cada SKU con stock_inicial+entrada (no hay lotes)
    n_inv_mov = 0
    for inv in inversores:
        pid = productos_ids.get(inv["codigo"])
        if not pid:
            continue
        cantidad_total = inv["stock_inicial"] + inv["entrada"]
        if cantidad_total <= 0:
            continue
        motivo = "Carga inicial inversores - existencia inicial + entradas del periodo"
        obs = (
            f"Inicial: {inv['stock_inicial']}, entrada: {inv['entrada']}, "
            f"salida: {inv['salida']}. Stock actual: {inv['stock']}. "
            f"Precio compra: {inv['precio_usd']:.2f} USD/u. "
            f"Pendiente reconciliación contable con control."
        )
        if insertar_movimiento(
            almacen_id, pid, cantidad_total, capturado_por, motivo, obs
        ):
            n_inv_mov += 1
        # Y si hubo salidas, registrarlas con tipo salida_ajuste y cantidad positiva
        if inv["salida"] > 0:
            obs_salida = (
                f"Salidas previas registradas en Excel original. "
                f"No se vinculan a proyecto/OT por falta de información."
            )
            insertar_movimiento(
                almacen_id, pid, inv["salida"], capturado_por,
                "Salida histórica - registro inicial sin proyecto destino",
                obs_salida,
                tipo="salida_ajuste",
                registro_contable_pendiente=False,  # Ya están fuera, no se reconcilia
            )
    print(f"  {n_inv_mov} movimientos iniciales de inversores.")

    n_cab = 0
    for cab in cables:
        pid = productos_ids.get(cab["codigo"])
        if not pid:
            continue
        if cab["stock_inicial"] > 0:
            if insertar_movimiento(
                almacen_id,
                pid,
                cab["stock_inicial"],
                capturado_por,
                "Carga inicial cables FV - existencia al iniciar registro",
                f"Stock inicial: {cab['stock_inicial']} {cab['unidad_medida']}. "
                f"Precio compra: {cab['precio_usd']:.4f} USD/u. "
                f"Pendiente reconciliación contable con control.",
            ):
                n_cab += 1
        if cab["salida"] > 0:
            insertar_movimiento(
                almacen_id, pid, cab["salida"], capturado_por,
                "Salida histórica cables - registro inicial sin proyecto",
                f"Salida de {cab['salida']} {cab['unidad_medida']} registrada en Excel original.",
                tipo="salida_ajuste",
                registro_contable_pendiente=False,
            )
    print(f"  {n_cab} movimientos iniciales de cables.")

    print("\n=== Resumen ===")
    print(f"  Productos en catálogo: {len(productos_ids)}")
    print(f"  Stock cargado        : {n_inv} SKUs con stock > 0")
    print(f"  Movimientos creados  : {n_mov + n_inv_mov + n_cab} aprox")
    print()
    print("✓ Importación completa.")
    print()
    print("Pendiente para CONTROL:")
    print("  - Llenar empresa_pago_id en cada movimiento según mapeo de color")
    print("  - Decidir reconciliación contable (capitalizar activo / dejar como gasto / etc.)")
    print("  - Query: SELECT * FROM v_inventario_pendientes_contables;")


if __name__ == "__main__":
    main()
