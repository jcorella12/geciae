# -*- coding: utf-8 -*-
"""
Repara el inventario importado en la primera corrida:

1. Corrige categoria en catalogo_productos (las primeras se guardaron como
   strings libres, deben ser valores del enum categoria_inventario:
   panel_solar, inversor, estructura, cable, herraje, monitoreo).
2. Agrega los movimientos de salida_ajuste que fallaron en la primera
   importación porque venían con cantidad negativa.

Uso:
    python scripts/reparar_inventario.py
"""

import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

SRC = Path(__file__).parent.parent / "data" / "inventario" / "inventario-consolidado.xlsx"

SUPABASE_URL = "https://dtmcqjtqykbkapzebbik.supabase.co"
SERVICE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6"
    "ImR0bWNxanRxeWtia2FwemViYmlrIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6"
    "MTc3NzU5MjM4MCwiZXhwIjoyMDkzMTY4MzgwfQ.-0TfmY0JaZTSn62jqtcmeroNeLX99"
    "Soa654OVbO1hwY"
)

H = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
}


def http(method: str, path: str, body=None, prefer=None) -> tuple[int, str]:
    h = dict(H)
    if prefer:
        h["Prefer"] = prefer
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{path}", data=data, method=method, headers=h
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.status, r.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8")


# -------- Mapeo de categoria_string_libre -> enum --------
CATEGORIA_FIX = {
    "panel": "panel_solar",
    "estructura_montaje": "estructura",
    "cable_solar": "cable",
    "conector_mc4": "herraje",
    "inversor_aislado": "inversor",
    "microinversor": "inversor",
    "accesorio": "herraje",
}


def reparar_categorias():
    print("=== 1. Reparando categorías ===")
    s, body = http("GET", "catalogo_productos?select=id,codigo,categoria")
    productos = json.loads(body)
    print(f"   {len(productos)} productos en catálogo.")
    n_fix = 0
    for p in productos:
        cat = p.get("categoria")
        if cat in CATEGORIA_FIX:
            nueva = CATEGORIA_FIX[cat]
            s2, body2 = http(
                "PATCH",
                f"catalogo_productos?id=eq.{p['id']}",
                body={"categoria": nueva},
                prefer="return=minimal",
            )
            if s2 in (200, 204):
                n_fix += 1
                print(f"   {p['codigo']:<25} {cat} → {nueva}")
            else:
                print(f"   ERROR fixing {p['codigo']}: {s2} {body2[:150]}")
    print(f"   {n_fix} productos actualizados.")


def encontrar_almacen_y_capturador() -> tuple[str, str]:
    s, body = http("GET", "almacenes?select=id&codigo=eq.CONTENEDORES&limit=1")
    almacen_id = json.loads(body)[0]["id"]
    s, body = http("GET", "usuarios_empresas?select=usuario_id&rol=eq.ceo&limit=1")
    cap = json.loads(body)[0]["usuario_id"]
    return almacen_id, cap


def reparar_salidas_inversores(wb, almacen_id: str, capturado_por: str):
    print()
    print("=== 2. Insertando salidas_ajuste de inversores ===")
    ws = wb["Inversores"]
    # Obtener IDs por codigo
    s, body = http("GET", "catalogo_productos?select=id,codigo")
    pid_by_codigo = {p["codigo"]: p["id"] for p in json.loads(body)}

    n_ok = 0
    n_skip = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        codigo, marca, descripcion, categoria, capacidad, ini, ent, sal, *_ = row
        if not codigo:
            continue
        salida = int(sal or 0)
        if salida <= 0:
            n_skip += 1
            continue
        pid = pid_by_codigo.get(str(codigo))
        if not pid:
            print(f"   No existe producto {codigo}, saltando salida")
            continue
        payload = {
            "almacen_id": almacen_id,
            "producto_id": pid,
            "tipo": "salida_ajuste",
            "cantidad": salida,
            "capturado_por": capturado_por,
            "motivo": "Salida histórica - registro inicial sin proyecto destino",
            "observaciones": (
                f"Salida de {salida} {descripcion} registrada en Excel original. "
                "Sin vínculo a proyecto/OT por falta de información en captura."
            )[:1000],
            "registro_contable_pendiente": False,
        }
        s2, body2 = http(
            "POST", "inventario_movimientos", body=payload, prefer="return=minimal"
        )
        if s2 in (200, 201, 204):
            n_ok += 1
            print(f"   {codigo:<30} salida {salida}")
        else:
            print(f"   ERROR salida {codigo}: {s2} {body2[:200]}")
    print(f"   {n_ok} salidas insertadas, {n_skip} sin salida.")


def reparar_salidas_cables(wb, almacen_id: str, capturado_por: str):
    print()
    print("=== 3. Insertando salidas_ajuste de cables ===")
    ws = wb["Cables FV"]
    s, body = http("GET", "catalogo_productos?select=id,codigo")
    pid_by_codigo = {p["codigo"]: p["id"] for p in json.loads(body)}
    n_ok = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        codigo, descripcion, unidad, ini, ent, sal, *_ = row
        if not codigo:
            continue
        salida = int(sal or 0)
        if salida <= 0:
            continue
        pid = pid_by_codigo.get(str(codigo))
        if not pid:
            continue
        payload = {
            "almacen_id": almacen_id,
            "producto_id": pid,
            "tipo": "salida_ajuste",
            "cantidad": salida,
            "capturado_por": capturado_por,
            "motivo": "Salida histórica cables - registro inicial sin proyecto",
            "observaciones": f"Salida de {salida} {unidad} de {descripcion}.",
            "registro_contable_pendiente": False,
        }
        s2, body2 = http(
            "POST", "inventario_movimientos", body=payload, prefer="return=minimal"
        )
        if s2 in (200, 201, 204):
            n_ok += 1
            print(f"   {codigo:<25} salida {salida} {unidad}")
        else:
            print(f"   ERROR salida {codigo}: {s2} {body2[:200]}")
    print(f"   {n_ok} salidas de cables insertadas.")


def main():
    if not SRC.exists():
        raise SystemExit(f"No existe {SRC}")
    print(f"Leyendo {SRC}...\n")
    wb = openpyxl.load_workbook(SRC, data_only=False)
    almacen_id, capturado_por = encontrar_almacen_y_capturador()

    reparar_categorias()
    reparar_salidas_inversores(wb, almacen_id, capturado_por)
    reparar_salidas_cables(wb, almacen_id, capturado_por)

    print("\n✓ Reparación completa.")


if __name__ == "__main__":
    main()
