"""
Sprint Z+ — Consolidar inventario crudo del cliente.

Lee `data/inventario/inventario-por-afinar-original.xlsx` y produce una versión
limpia que:

1. **Preserva los lotes originales** en hoja "Lotes detalle" con etiqueta de
   color → empresa (cada compra del mismo SKU vino de una empresa distinta
   pero todo se acumula al inventario del grupo).
2. **Consolida por SKU** sumando cantidades (precio promedio ponderado) en las
   hojas "Estructura PV", "Inversores", "Cables FV".
3. **TC editable**: hoja "Configuración" tiene el tipo de cambio USD→MXN en
   `$B$1`. Todas las columnas MXN son fórmulas `=USD * Configuración!$B$1`.
   Cambiar la celda recalcula todo el archivo.
4. Corrige typo "NXT-TBM8x25-5" → "NXT-TB-M8X25-5".
5. Genera código único por marca+modelo para inversores (antes "CODIGO" era
   solo la marca).

Uso:
    python scripts/limpiar_inventario.py
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SRC = PROJECT_ROOT / "data" / "inventario" / "inventario-por-afinar-original.xlsx"
DST = PROJECT_ROOT / "data" / "inventario" / "inventario-consolidado.xlsx"

TC_DEFAULT = 17.2530  # USD/MXN FIX más reciente (Banxico SF43718, 2026-05-06)

# --- Correcciones del cliente ---
# Cuando el cliente reporta errores en el archivo original, se documentan aquí.
# Se aplican durante la consolidación pero NO modifican el archivo original
# (auditoría preservada en hoja "Lotes detalle").
#
# Formato: codigo_canonico → { campo: valor_corregido, motivo: str }
CORRECCIONES_LOTES: dict[str, dict] = {
    "JA-M66D45-620/LB": {
        "cantidad": 1546,
        "motivo": "Cliente confirma 1546 paneles, no 15 (error de captura original)",
    },
}

# --- Estilos ---
HEADER_FILL = PatternFill("solid", start_color="1A1A2E")
HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(name="Arial", size=10)
TOTAL_FONT = Font(name="Arial", bold=True, size=11)
TOTAL_FILL = PatternFill("solid", start_color="EAEAEA")
CONFIG_FILL = PatternFill("solid", start_color="FFF8DC")
THIN = Side(border_style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- Mapeo color → label provisional ---
# El usuario llenará la columna empresa en la hoja Configuración.
# Los colores del theme se observaron en la inspección del original.
COLOR_LABELS = {
    "theme:9": "Color A (lote más frecuente)",
    "theme:8": "Color B",
    "FFFFFF00": "Color C (amarillo)",
}


def normalizar_codigo(codigo: str) -> str:
    return (codigo or "").strip().upper().replace("  ", " ")


def codigo_canonico(c: str) -> str:
    """Resuelve typos conocidos a códigos canónicos."""
    c = normalizar_codigo(c)
    fixes = {
        "NXT-TBM8X25-5": "NXT-TB-M8X25-5",
    }
    return fixes.get(c, c)


def detectar_color(cell: Any) -> str:
    """Devuelve un identificador estable del color de fondo."""
    fill = cell.fill
    if not fill or not fill.fgColor:
        return "NONE"
    fg = fill.fgColor
    if fg.type == "rgb" and fg.rgb:
        return str(fg.rgb).upper()
    if fg.type == "theme":
        return f"theme:{fg.theme}"
    return "NONE"


def leer_lotes_estructura(wb_src: openpyxl.Workbook) -> list[dict]:
    """Lee Hoja1 preservando cada compra (lote) original con su color."""
    sheet = wb_src["Hoja1"]
    lotes: list[dict] = []
    for row_idx in range(3, sheet.max_row + 1):
        codigo_cell = sheet.cell(row=row_idx, column=1)
        codigo = codigo_cell.value
        if not codigo:
            continue
        color = detectar_color(codigo_cell)
        descripcion = sheet.cell(row=row_idx, column=2).value or ""
        exist_ini = sheet.cell(row=row_idx, column=3).value or 0
        entrada = sheet.cell(row=row_idx, column=4).value or 0
        salida = sheet.cell(row=row_idx, column=5).value or 0
        precio = sheet.cell(row=row_idx, column=7).value or 0
        cantidad = int(entrada) - int(salida) + int(exist_ini)
        cod_canon = codigo_canonico(str(codigo))

        # Aplicar correcciones del cliente (documentadas arriba en CORRECCIONES_LOTES)
        correccion = CORRECCIONES_LOTES.get(cod_canon)
        cantidad_original = cantidad
        if correccion and "cantidad" in correccion:
            cantidad = int(correccion["cantidad"])

        if cantidad <= 0:
            continue
        lotes.append(
            {
                "codigo": cod_canon,
                "descripcion": str(descripcion).strip(),
                "cantidad": cantidad,
                "cantidad_original": cantidad_original,
                "precio_unit_usd": float(precio),
                "color": color,
                "color_label": COLOR_LABELS.get(color, f"Otro ({color})"),
                "row_origen": row_idx,
                "corregido": correccion is not None,
                "motivo_correccion": (correccion or {}).get("motivo"),
            }
        )
    return lotes


def consolidar_lotes(lotes: list[dict]) -> list[dict]:
    """Agrupa lotes por código y calcula precio promedio ponderado."""
    grupos: dict[str, dict] = defaultdict(
        lambda: {
            "descripcion": "",
            "cantidad": 0,
            "valor_usd": 0.0,
            "lotes": 0,
            "colores": set(),
            "corregido": False,
        }
    )
    for l in lotes:
        g = grupos[l["codigo"]]
        if not g["descripcion"]:
            g["descripcion"] = l["descripcion"]
        g["cantidad"] += l["cantidad"]
        g["valor_usd"] += l["cantidad"] * l["precio_unit_usd"]
        g["lotes"] += 1
        g["colores"].add(l["color"])
        if l.get("corregido"):
            g["corregido"] = True

    items = []
    for codigo, g in grupos.items():
        precio_prom = g["valor_usd"] / g["cantidad"] if g["cantidad"] else 0
        items.append(
            {
                "codigo": codigo,
                "descripcion": g["descripcion"],
                "stock": g["cantidad"],
                "precio_unit_usd": round(precio_prom, 4),
                "valor_total_usd": round(g["valor_usd"], 2),
                "n_lotes": g["lotes"],
                "n_empresas": len(g["colores"]),
                "corregido": g["corregido"],
            }
        )
    items.sort(key=lambda x: x["codigo"])
    return items


def categorizar_inversor(descripcion: str) -> str:
    d = (descripcion or "").upper()
    if "MICRO" in d:
        return "microinversor"
    if "AISLADO" in d:
        return "inversor_aislado"
    if "ECU" in d or "MONITOREO" in d:
        return "monitoreo"
    if "CABLE" in d or "ENDCAP" in d or "BUS" in d:
        return "accesorio"
    return "inversor"


def codigo_inversor(marca: str, descripcion: str) -> str:
    """Genera código único combinando marca + modelo extraído."""
    marca_corta = re.sub(r"[^A-Z]", "", (marca or "").upper())[:6] or "INV"
    desc = re.sub(rf"^{re.escape(marca or '')}\s*", "", descripcion or "", flags=re.I)
    modelo = re.sub(r"[^A-Z0-9.]", "", desc.upper())[:20]
    if not modelo:
        modelo = "MODELO"
    return f"{marca_corta}-{modelo}"


def leer_inversores_y_cables(wb_src: openpyxl.Workbook) -> tuple[list[dict], list[dict]]:
    sheet = wb_src["Sheet1"]
    inversores = []
    cables = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        marca, descripcion, potencia, exist_ini, entrada, salida, stok, *resto = row
        if not marca and not descripcion:
            continue
        if marca and "CABLEADO" in str(marca).upper():
            cables.append(
                {
                    "codigo": str(marca).strip().replace(" ", "-"),
                    "descripcion": (descripcion or "").strip(),
                    "unidad": "metros",
                    "stock_inicial": int(exist_ini or 0),
                    "entrada": int(entrada or 0),
                    "salida": int(salida or 0),
                    "stock_actual": int(stok or 0),
                    "precio_unit_usd": float(resto[1] or 0),
                }
            )
            continue
        if not marca or not descripcion:
            continue
        codigo = codigo_inversor(marca, descripcion)
        precio_w = float(resto[0] or 0) if resto else 0
        precio_unit = float(resto[1] or 0) if len(resto) > 1 else 0
        if isinstance(potencia, (int, float)):
            capacidad_w = float(potencia)
            cap_label = f"{capacidad_w / 1000:g} kW"
        else:
            capacidad_w = None
            cap_label = str(potencia or "")
        inversores.append(
            {
                "codigo": codigo,
                "marca": (marca or "").strip(),
                "descripcion": (descripcion or "").strip(),
                "categoria": categorizar_inversor(descripcion or ""),
                "capacidad_w": capacidad_w,
                "capacidad_label": cap_label,
                "stock_inicial": int(exist_ini or 0) if exist_ini is not None else 0,
                "entrada": int(entrada or 0),
                "salida": int(salida or 0),
                "stock_actual": int(stok or 0),
                "precio_x_watt_usd": round(precio_w, 4) if precio_w else None,
                "precio_unit_usd": round(precio_unit, 2),
            }
        )
    inversores.sort(key=lambda x: (x["marca"], x["codigo"]))
    return inversores, cables


# ---------------------------------------------------------------------------
# Escritura
# ---------------------------------------------------------------------------


def header(ws, fila: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=fila, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[fila].height = 32


def autoancho(ws, columnas: int, max_w: int = 50) -> None:
    for c in range(1, columnas + 1):
        max_len = 0
        for cell in ws[get_column_letter(c)]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(c)].width = min(max(12, max_len + 2), max_w)


def escribir_configuracion(wb: openpyxl.Workbook, colores_usados: set[str]) -> None:
    ws = wb.create_sheet("Configuración")

    # Título
    ws.cell(row=1, column=1, value="Tipo de cambio USD→MXN").font = Font(
        name="Arial", bold=True, size=12
    )
    tc_cell = ws.cell(row=1, column=2, value=TC_DEFAULT)
    tc_cell.number_format = "$#,##0.0000"
    tc_cell.fill = CONFIG_FILL
    tc_cell.font = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    tc_cell.border = BORDER
    tc_cell.alignment = Alignment(horizontal="center")

    ws.cell(
        row=2,
        column=1,
        value="↑ Edita esta celda y todos los valores MXN del archivo se recalculan",
    ).font = Font(name="Arial", italic=True, size=10, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    # Nombre definido para que las fórmulas usen `tc_actual` en vez de la celda
    wb.defined_names.add(DefinedName("tc_actual", attr_text="Configuración!$B$1"))

    # Título mapeo de colores
    ws.cell(row=4, column=1, value="Mapeo de color → empresa").font = Font(
        name="Arial", bold=True, size=12
    )
    ws.cell(
        row=5,
        column=1,
        value=(
            "En el Excel original cada lote tenía un color que indica de qué empresa "
            "salió el gasto. Llena la columna Empresa para que quede registrado:"
        ),
    ).font = Font(name="Arial", italic=True, size=10, color="666666")
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)

    header(ws, 7, ["Color", "Etiqueta", "Empresa (llena tú)", "Notas"])
    fila = 8
    for color in sorted(colores_usados):
        ws.cell(row=fila, column=1, value=color)
        ws.cell(row=fila, column=2, value=COLOR_LABELS.get(color, f"Otro ({color})"))
        ws.cell(row=fila, column=3, value="").fill = CONFIG_FILL
        ws.cell(row=fila, column=4, value="(PSE / CIAE / IED / Limson)").font = Font(
            name="Arial", italic=True, size=9, color="888888"
        )
        for c in range(1, 5):
            ws.cell(row=fila, column=c).border = BORDER
        fila += 1

    # Notas generales
    fila += 2
    ws.cell(row=fila, column=1, value="Notas").font = Font(name="Arial", bold=True, size=12)
    fila += 1
    notas = [
        "• Los costos de este archivo NO son los de hoy — son históricos al momento de la compra.",
        "• El TC en B1 actualiza todos los valores MXN del archivo (fórmulas dinámicas).",
        "• Hoja1 original tenía 38 filas con varias compras del mismo SKU. Cada compra venía de",
        "  una empresa distinta (color), pero TODO se acumula al inventario del grupo porque ya",
        "  se contabilizó como gasto en cada empresa.",
        "• Sheet1 original usaba la marca como CODIGO. Se generaron códigos únicos combinando",
        '  marca + modelo (ej: "GROWAT-MAC15KTL3").',
        "• Inversores con stock=0 se mantienen en el catálogo para histórico de movimientos.",
        "• Hoja 'Lotes detalle' preserva cada compra original con su color para auditoría.",
        "• Typo corregido: NXT-TBM8x25-5 → NXT-TB-M8X25-5.",
    ]
    for n in notas:
        ws.cell(row=fila, column=1, value=n).font = Font(name="Arial", size=10)
        fila += 1

    autoancho(ws, 4, max_w=80)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25


def escribir_lotes_detalle(wb: openpyxl.Workbook, lotes: list[dict]) -> None:
    ws = wb.create_sheet("Lotes detalle")

    headers = [
        "Código",
        "Descripción",
        "Cantidad",
        "Cant. orig. (Excel)",
        "Corrección",
        "Precio unit. USD",
        "Color (origen)",
        "Etiqueta color",
        "Valor lote USD",
        "Valor lote MXN",
    ]
    header(ws, 1, headers)

    correccion_fill = PatternFill("solid", start_color="FFF3CD")  # Amarillo suave

    fila = 2
    for l in lotes:
        ws.cell(row=fila, column=1, value=l["codigo"])
        ws.cell(row=fila, column=2, value=l["descripcion"])
        ws.cell(row=fila, column=3, value=l["cantidad"]).number_format = "#,##0"
        # Mostrar cantidad original solo si difiere
        if l.get("corregido"):
            ws.cell(row=fila, column=4, value=l["cantidad_original"]).number_format = "#,##0"
            ws.cell(row=fila, column=5, value=l.get("motivo_correccion") or "Corregido")
            for c in range(3, 6):
                ws.cell(row=fila, column=c).fill = correccion_fill
        else:
            ws.cell(row=fila, column=4, value="—")
            ws.cell(row=fila, column=5, value="")
        ws.cell(row=fila, column=6, value=l["precio_unit_usd"]).number_format = "$#,##0.0000"
        ws.cell(row=fila, column=7, value=l["color"])
        ws.cell(row=fila, column=8, value=l["color_label"])
        ws.cell(row=fila, column=9, value=f"=C{fila}*F{fila}").number_format = "$#,##0.00"
        ws.cell(row=fila, column=10, value=f"=I{fila}*tc_actual").number_format = "$#,##0.00"
        for c in range(1, len(headers) + 1):
            ws.cell(row=fila, column=c).font = BODY_FONT
            ws.cell(row=fila, column=c).border = BORDER
        fila += 1

    # Total
    ws.cell(row=fila, column=2, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=fila, column=9, value=f"=SUM(I2:I{fila - 1})").number_format = "$#,##0.00"
    ws.cell(row=fila, column=10, value=f"=SUM(J2:J{fila - 1})").number_format = "$#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=fila, column=c).fill = TOTAL_FILL
        ws.cell(row=fila, column=c).font = TOTAL_FONT

    autoancho(ws, len(headers))
    ws.freeze_panes = "A2"


def escribir_estructura(wb: openpyxl.Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Estructura PV")

    headers = [
        "Código",
        "Descripción",
        "Stock",
        "# Lotes",
        "# Empresas",
        "Precio prom. USD",
        "Valor total USD",
        "Valor total MXN",
    ]
    header(ws, 1, headers)

    correccion_fill = PatternFill("solid", start_color="FFF3CD")

    fila = 2
    for it in items:
        ws.cell(row=fila, column=1, value=it["codigo"])
        ws.cell(row=fila, column=2, value=it["descripcion"])
        stock_cell = ws.cell(row=fila, column=3, value=it["stock"])
        stock_cell.number_format = "#,##0"
        if it.get("corregido"):
            stock_cell.fill = correccion_fill
            stock_cell.comment = openpyxl.comments.Comment(
                "Stock corregido por instrucción del cliente. Ver hoja 'Lotes detalle'.",
                "Sistema",
            )
        ws.cell(row=fila, column=4, value=it["n_lotes"]).number_format = "#,##0"
        ws.cell(row=fila, column=5, value=it["n_empresas"]).number_format = "#,##0"
        ws.cell(row=fila, column=6, value=it["precio_unit_usd"]).number_format = "$#,##0.0000"
        ws.cell(row=fila, column=7, value=f"=C{fila}*F{fila}").number_format = "$#,##0.00"
        ws.cell(row=fila, column=8, value=f"=G{fila}*tc_actual").number_format = "$#,##0.00"
        for c in range(1, len(headers) + 1):
            ws.cell(row=fila, column=c).font = BODY_FONT
            ws.cell(row=fila, column=c).border = BORDER
        fila += 1

    ws.cell(row=fila, column=2, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=fila, column=7, value=f"=SUM(G2:G{fila - 1})").number_format = "$#,##0.00"
    ws.cell(row=fila, column=8, value=f"=SUM(H2:H{fila - 1})").number_format = "$#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=fila, column=c).fill = TOTAL_FILL
        ws.cell(row=fila, column=c).font = TOTAL_FONT

    autoancho(ws, len(headers))
    ws.freeze_panes = "A2"


def escribir_inversores(wb: openpyxl.Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Inversores")

    headers = [
        "Código",
        "Marca",
        "Descripción",
        "Categoría",
        "Capacidad",
        "Stock inicial",
        "Entrada",
        "Salida",
        "Stock actual",
        "Precio/W USD",
        "Precio unit. USD",
        "Valor total USD",
        "Valor total MXN",
    ]
    header(ws, 1, headers)

    fila = 2
    for it in items:
        ws.cell(row=fila, column=1, value=it["codigo"])
        ws.cell(row=fila, column=2, value=it["marca"])
        ws.cell(row=fila, column=3, value=it["descripcion"])
        ws.cell(row=fila, column=4, value=it["categoria"])
        ws.cell(row=fila, column=5, value=it["capacidad_label"])
        ws.cell(row=fila, column=6, value=it["stock_inicial"]).number_format = "#,##0"
        ws.cell(row=fila, column=7, value=it["entrada"]).number_format = "#,##0"
        ws.cell(row=fila, column=8, value=it["salida"]).number_format = "#,##0"
        ws.cell(row=fila, column=9, value=f"=F{fila}+G{fila}-H{fila}").number_format = "#,##0"
        ws.cell(row=fila, column=10, value=it["precio_x_watt_usd"]).number_format = "$#,##0.0000"
        ws.cell(row=fila, column=11, value=it["precio_unit_usd"]).number_format = "$#,##0.00"
        ws.cell(row=fila, column=12, value=f"=I{fila}*K{fila}").number_format = "$#,##0.00"
        ws.cell(row=fila, column=13, value=f"=L{fila}*tc_actual").number_format = "$#,##0.00"
        for c in range(1, len(headers) + 1):
            ws.cell(row=fila, column=c).font = BODY_FONT
            ws.cell(row=fila, column=c).border = BORDER
        fila += 1

    ws.cell(row=fila, column=2, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=fila, column=12, value=f"=SUM(L2:L{fila - 1})").number_format = "$#,##0.00"
    ws.cell(row=fila, column=13, value=f"=SUM(M2:M{fila - 1})").number_format = "$#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=fila, column=c).fill = TOTAL_FILL
        ws.cell(row=fila, column=c).font = TOTAL_FONT

    autoancho(ws, len(headers))
    ws.freeze_panes = "A2"


def escribir_cables(wb: openpyxl.Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Cables FV")

    headers = [
        "Código",
        "Descripción",
        "Unidad",
        "Stock inicial",
        "Entrada",
        "Salida",
        "Stock actual",
        "Precio unit. USD",
        "Valor total USD",
        "Valor total MXN",
    ]
    header(ws, 1, headers)

    fila = 2
    for it in items:
        ws.cell(row=fila, column=1, value=it["codigo"])
        ws.cell(row=fila, column=2, value=it["descripcion"])
        ws.cell(row=fila, column=3, value=it["unidad"])
        ws.cell(row=fila, column=4, value=it["stock_inicial"]).number_format = "#,##0"
        ws.cell(row=fila, column=5, value=it["entrada"]).number_format = "#,##0"
        ws.cell(row=fila, column=6, value=it["salida"]).number_format = "#,##0"
        ws.cell(row=fila, column=7, value=f"=D{fila}+E{fila}-F{fila}").number_format = "#,##0"
        ws.cell(row=fila, column=8, value=it["precio_unit_usd"]).number_format = "$#,##0.0000"
        ws.cell(row=fila, column=9, value=f"=G{fila}*H{fila}").number_format = "$#,##0.00"
        ws.cell(row=fila, column=10, value=f"=I{fila}*tc_actual").number_format = "$#,##0.00"
        for c in range(1, len(headers) + 1):
            ws.cell(row=fila, column=c).font = BODY_FONT
            ws.cell(row=fila, column=c).border = BORDER
        fila += 1

    ws.cell(row=fila, column=2, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=fila, column=9, value=f"=SUM(I2:I{fila - 1})").number_format = "$#,##0.00"
    ws.cell(row=fila, column=10, value=f"=SUM(J2:J{fila - 1})").number_format = "$#,##0.00"
    for c in range(1, len(headers) + 1):
        ws.cell(row=fila, column=c).fill = TOTAL_FILL
        ws.cell(row=fila, column=c).font = TOTAL_FONT

    autoancho(ws, len(headers))
    ws.freeze_panes = "A2"


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"No existe {SRC}")
    print(f"Leyendo {SRC}...")
    wb_src = openpyxl.load_workbook(SRC, data_only=True)

    lotes = leer_lotes_estructura(wb_src)
    estructura = consolidar_lotes(lotes)
    inversores, cables = leer_inversores_y_cables(wb_src)

    colores_usados = set(l["color"] for l in lotes)

    print(f"  Lotes en Hoja1            : {len(lotes)} (de {sum(1 for l in lotes if l)} compras)")
    print(f"  SKUs únicos estructura    : {len(estructura)}")
    print(f"  Colores distintos detec.  : {len(colores_usados)} → {sorted(colores_usados)}")
    print(f"  Inversores únicos         : {len(inversores)}")
    print(f"  Cables sueltos            : {len(cables)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Orden: Configuración → resúmenes → detalle de lotes
    escribir_configuracion(wb, colores_usados)
    escribir_estructura(wb, estructura)
    escribir_inversores(wb, inversores)
    escribir_cables(wb, cables)
    escribir_lotes_detalle(wb, lotes)

    wb.save(DST)
    print(f"\nGuardado en {DST}")
    print(f"  Hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
